import io
import time
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_chat_table(df, title="Deals Preview", top_n=5, cols_to_show=None):
    if df.empty: return ""
    cols = cols_to_show if cols_to_show else ['Symbol', 'Buy/Sell', 'QuantityTraded', 'TradePrice/Wght.Avg.Price']
    existing_cols = [c for c in cols if c in df.columns]
    
    preview_df = df[existing_cols].head(top_n).copy()
    if 'QuantityTraded' in preview_df.columns:
        preview_df['QuantityTraded'] = preview_df['QuantityTraded'].apply(lambda x: f"{x:,}")
    if 'TradePrice/Wght.Avg.Price' in preview_df.columns:
        preview_df['TradePrice/Wght.Avg.Price'] = preview_df['TradePrice/Wght.Avg.Price'].apply(lambda x: f"₹{x:,.2f}")
        
    col_rename = {'Symbol': 'Symbol', 'Ticker': 'Ticker', 'Buy/Sell': 'Type', 'QuantityTraded': 'Qty', 'TradePrice/Wght.Avg.Price': 'Price', 'Entry Price': 'Entry', 'Exit Price': 'Exit', 'PnL (%)': 'PnL%'}
    preview_df.rename(columns=col_rename, inplace=True)
    return f"📊 *{title} (Top {len(preview_df)} Records)*\n```\n{preview_df.to_string(index=False)}\n```"

def format_excel_sheet(ws, df_sheet, title_text, default_days_or_info):
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=15, bold=True, color="1B365D")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="333333")
    
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    thin_border = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'), top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))

    ws.views.sheetView[0].showGridLines = True
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"{title_text} ({default_days_or_info})"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 10
    
    for col_idx, h in enumerate(list(df_sheet.columns), 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = str(h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    ws.row_dimensions[3].height = 25
    row_idx = 4
    for _, row in df_sheet.iterrows():
        ws.row_dimensions[row_idx].height = 20
        row_fill = zebra_fill if row_idx % 2 == 0 else white_fill
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = str(val) if not isinstance(val, (int, float)) else val
            cell.font = data_font
            cell.border = thin_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="left" if isinstance(val, str) else "right", vertical="center")
        row_idx += 1
        
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 18

def process_df(df):
    if df.empty: return df
    df.columns = [c.strip() for c in df.columns]
    
    def parse_date(d_str):
        for fmt in ("%d-%b-%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try: return datetime.strptime(str(d_str).strip(), fmt)
            except ValueError: continue
        return str(d_str).strip()
        
    df['ParsedDate'] = df['Date'].apply(parse_date)
    df = df.sort_values(by=['ParsedDate', 'Symbol'], ascending=[False, True]) if pd.api.types.is_datetime64_any_dtype(df['ParsedDate']) else df.sort_values(by=['Date', 'Symbol'], ascending=[False, True])
    if pd.api.types.is_datetime64_any_dtype(df['ParsedDate']): df['Date'] = df['ParsedDate'].dt.strftime('%d-%b-%y')
    df = df.drop(columns=['ParsedDate'])
    df['QuantityTraded'] = pd.to_numeric(df['QuantityTraded'].astype(str).str.replace(',', ''), errors='coerce').fillna(0).astype(int)
    df['TradePrice/Wght.Avg.Price'] = pd.to_numeric(df['TradePrice/Wght.Avg.Price'].astype(str).str.replace(',', ''), errors='coerce').fillna(0.0)
    df['TradeValue_INR'] = df['QuantityTraded'] * df['TradePrice/Wght.Avg.Price']
    df['Buy/Sell'] = df['Buy/Sell'].astype(str).str.strip().str.upper()
    return df

def get_excel_buffer(df_bulk, df_block, title_prefix, sub_title):
    wb = openpyxl.Workbook()
    wb.remove(wb.active) 
    if not df_bulk.empty: format_excel_sheet(wb.create_sheet("Bulk Deals"), df_bulk, f"{title_prefix} Bulk Deals", sub_title)
    if not df_block.empty: format_excel_sheet(wb.create_sheet("Block Deals"), df_block, f"{title_prefix} Block Deals", sub_title)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def get_sma_custom_excel_buffer(df_dict):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, color="1F2937")
    profit_font = Font(name="Segoe UI", size=10, bold=True, color="00C04B")
    loss_font = Font(name="Segoe UI", size=10, bold=True, color="FF3131")
    link_font = Font(name="Segoe UI", size=10, color="3B82F6", underline="single")

    for sheet_name, df in df_dict.items():
        if df.empty: continue
        ws = wb.create_sheet(sheet_name)
        headers = list(df.columns)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(h))
            cell.font, cell.fill, cell.alignment = header_font, header_fill, Alignment(horizontal="center")
            
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, val in enumerate(row, 1):
                col_name = headers[col_idx-1]
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_name in ["Screener", "TradingView"]:
                    cell.value, cell.hyperlink, cell.font, cell.alignment = col_name, val, link_font, Alignment(horizontal="center")
                elif col_name == "PnL (%)":
                    cell.value = val / 100 if isinstance(val, (int, float)) else val
                    cell.number_format = '+0.00%;-0.00%'
                    if isinstance(val, (int, float)): cell.font = profit_font if val > 0 else loss_font
                    cell.alignment = Alignment(horizontal="right")
                elif "Price" in col_name:
                    cell.value, cell.number_format, cell.font = val, '#,##0.00', data_font
                else:
                    cell.value, cell.font, cell.alignment = val, data_font, Alignment(horizontal="center" if "Date" in col_name or "Days" in col_name else "left")

        for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width = 16
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer